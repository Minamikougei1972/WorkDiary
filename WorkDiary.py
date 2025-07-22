#-------やっとわかってきた！-toiunohausoda------
from __future__ import annotations
import os
import shutil
import sqlite3
import openpyxl
import tkinter as tk
from tkinter import messagebox
import datetime as dt
from pathlib import Path
import re
from copy import copy
import json
from pathlib import Path
from typing import List, Dict, Optional, Iterator


# ---- 定数 ----
HEADER_ROWS = 2                # 夜勤ヘッダーはテンプレートから 2 行コピー
ARTICLE_ROWS_PER_PAGE = 54     # 印刷 1 ページあたりの行数 (A4)
MAX_EMPTY_ROWS = 30            # 空行が続けばファイル終端とみなす


# ----------------- テンプレ裏シートの行数設定 -----------------
SHEET_TOTAL_ROWS        = 37   # 行番号 1〜37 を使用
SHEET_HEADER_ROWS       = 1    # 1 行目は常に見出し
ARTICLE_ROWS_PER_SHEET  = SHEET_TOTAL_ROWS - SHEET_HEADER_ROWS  # = 36
MIN_NIGHT_ROWS   = 5   # ヘッダー2行＋最低3行の本文を書きたい

# ------------------------------------------------------------------
# Part A : personal_pointer テーブルと基本ユーティリティ
# ------------------------------------------------------------------

def init_personal_tables(db_path: str):
    """
    personal_pointer テーブル（個人ファイルの書き込み位置管理）を作成する。
    residents テーブルは既に作成済みなのでここでは作らない。
    """
    conn = sqlite3.connect(db_path)
    cur  = conn.cursor()

    # name = 入所者氏名
    # file = 個人ファイル名 (2階 / 3階 / 退職者)
    # sheet = シート名 例『宮本武蔵』, 『宮本武蔵(2)』
    # next_row = 次に書き込む行番号 (ヘッダ 4 行目開始なので 4 以上)
    cur.execute('''
        CREATE TABLE IF NOT EXISTS personal_pointer (
            name      TEXT PRIMARY KEY,
            file      TEXT NOT NULL,
            sheet     TEXT NOT NULL,
            next_row  INTEGER NOT NULL
        )
    ''')
    conn.commit()
    conn.close()

# ----------------------------------------------------------------------------
#  ユーティリティ (純粋関数)
# ----------------------------------------------------------------------------

def normalize_text(val: str | None) -> str:
    """
    Noneの場合は空文字にし、前後の空白を除去して返す。
    Excelセル値の安定化用。
    """
    return val.strip() if val else ""


def iter_rows(sheet, start: int = 2) -> Iterator[tuple[int, str, str]]:
    """
    指定行からA列（名前）・B列（本文）の値を順に返すイテレータ。
    """
    row = start
    while row <= sheet.max_row:
        name = sheet.cell(row, 1).value
        content = sheet.cell(row, 2).value
        yield row, normalize_text(name), normalize_text(content)
        row += 1
# ----------------------------------------------------------------------------
#  1) データ取得  -------------------------------------------------------------
# ----------------------------------------------------------------------------

def extract_entries(sheet, *, row_start: int = 2) -> List[Dict]:
    """
    シート1枚から [{name, content, shift}] のリストを抽出する。
    DB書き込みやヘッダー貼付は行わない（純粋関数）。
    "以上"や"巡回"/"夜間浴"で日勤→夜勤の切替えを自動判定。
    """
    entries: List[Dict] = []
    current_name: Optional[str] = None
    current_content: List[str] = []
    current_shift = "日勤"
    empty_cnt = 0

    def flush():
        nonlocal current_name, current_content
        if current_name and current_content:
            entries.append({
                "name": current_name,
                "content": "\n".join(current_content),
                "shift": current_shift,
            })
        current_name, current_content = None, []

    for _, name, content in iter_rows(sheet, row_start):
        # 空行カウントで早期終了
        if not name and not content:
            empty_cnt += 1
            if empty_cnt >= MAX_EMPTY_ROWS:
                break
            continue
        empty_cnt = 0

        # "以上" 行で夜勤へ切替え
        if name == "以上":
            flush()
            current_shift = "夜勤"
            continue

        # 巡回/夜間浴 行はヘッダー扱い
        if name in ("巡回", "夜間浴"):
            if current_shift == "日勤":
                flush()
                current_shift = "夜勤"
            continue

        # 新利用者 or 同一利用者追記
        if name:
            flush()
            current_name = name
            current_content = [content] if content else []
        else:
            if content:
                current_content.append(content)

    flush()
    return entries


# ----------------------------------------------------------------------------
#  2) author 付与 (純粋関数)
# ----------------------------------------------------------------------------

def add_authors(entries: List[Dict], *, author_day: str, author_night: str) -> List[Dict]:
    """
    各エントリに shift（日勤/夜勤）ごとに author（日勤担当/夜勤担当）を付与して新リストを返す。
    """
    new_entries: List[Dict] = []
    for e in entries:
        e2 = e.copy()
        e2["author"] = author_day if e["shift"] == "日勤" else author_night
        new_entries.append(e2)
    return new_entries


# ----------------------------------------------------------------------------
#  3) DB 書き込み (副作用あり)
# ----------------------------------------------------------------------------

def save_entries_to_db(entries: List[Dict], db_path: str | Path, *, date: dt.date):
    """
    entriesの内容を日誌DB（diary_entriesテーブル）に保存する。
    既存重複は無視（INSERT OR IGNORE）。
    """
    conn = sqlite3.connect(db_path)
    cur = conn.cursor()
    date_str = date.strftime("%Y-%m-%d")

    rows = [(
        e["name"],
        date_str,
        e["shift"],
        e["content"],
        e["author"],
    ) for e in entries]

    cur.executemany(
        """INSERT OR IGNORE INTO diary_entries
           (resident_name, date, shift, content, author)
           VALUES (?, ?, ?, ?, ?)""",
        rows,
    )
    conn.commit()
    conn.close()


# ----------------------------------------------------------------------------
#  4) 見た目の更新 (副作用: Excel シートを書き換え)
# ----------------------------------------------------------------------------

def apply_night_header(sheet, row: int, template_sheet):
    """
    指定行(row)に夜勤ヘッダー（template_sheetの先頭2行）を貼り付ける。
    """
    if template_sheet is None:
        return
    for offset in range(HEADER_ROWS):
        for col in range(1, 3):
            dst = sheet.cell(row + offset, col)
            src = template_sheet.cell(offset + 1, col)
            dst.value = src.value
            if src.has_style:
                dst._style = src._style


def setup_page_breaks(sheet, rows_per_page: int = ARTICLE_ROWS_PER_PAGE):
    """
    指定行数ごとにExcelシートへ改ページを自動挿入する。
    """
    from openpyxl.worksheet.pagebreak import Break

    sheet.page_breaks = []  # 既存クリア
    idx = rows_per_page + 1
    while idx < sheet.max_row:
        sheet.page_breaks.append(Break(id=idx))
        idx += rows_per_page


def update_diary_sheet(sheet, *, template_sheet=None):
    """
    夜勤ヘッダー貼付と改ページ設定のみを行う。
    """
    current_shift = "日勤"
    for row_idx, name, _ in iter_rows(sheet, start=2):
        if name == "以上":
            apply_night_header(sheet, row_idx, template_sheet)
            current_shift = "夜勤"
            continue
        if name in ("巡回", "夜間浴") and current_shift == "日勤":
            current_shift = "夜勤"  # 既にヘッダーが入力済みなら何もしない
    setup_page_breaks(sheet)



# ---------- 便利関数群 -------------------------------------------------

def copy_left_of(wb, base_ws, template_name, new_title):
    """
    指定テンプレートシート(template_name)を複製し、base_wsの左隣にnew_titleで挿入。
    テンプレートがなければ空シートを作成。
    戻り値: 新しい Worksheet
    """
    if template_name in wb.sheetnames:
        new_ws = wb.copy_worksheet(wb[template_name])
    else:
        new_ws = wb.create_sheet()

    new_ws.title = new_title

    # ――― 位置を調整 ―――
    base_pos = wb.sheetnames.index(base_ws.title)
    wb._sheets.remove(new_ws)
    wb._sheets.insert(base_pos, new_ws)
    # ―――――――――――――――――

    return new_ws



def paste_night_header(sheet, row, tpl):
    """
    Header_Night(A1:B2) を sheet の row 行目に上書きコピーする。
    書式も含めてコピー。
    """
    for r in range(0, 2):          # 2 行
        for c in range(1, 3):      # 列 A,B
            src = tpl.cell(r + 1, c)
            dst = sheet.cell(row + r, c)
            dst.value         = src.value
            dst.font          = copy(src.font)
            dst.border        = copy(src.border)
            dst.fill          = copy(src.fill)
            dst.number_format = src.number_format
            dst.alignment     = copy(src.alignment)
            dst.protection    = copy(src.protection)



def add_ura_if_needed(file_path: str, base_sheet: str) -> None:
    """
    「○日裏(2)…」という名前のSheetを、
    ・まだ存在しなければB_tempからコピーして作成
    ・base_sheetの左（インデックス直前）に挿入
    """
    wb = openpyxl.load_workbook(file_path)

    # 例: base_sheet="15日裏" → base="15日裏"
    base, *_ = base_sheet.split("(")      # 「(」が無いときはそのまま
    idx = 2
    while f"{base}({idx})" in wb.sheetnames:
        idx += 1
    new_name = f"{base}({idx})"

    # まだ無く、テンプレート B_temp があるときだけ作成
    if "B_temp" in wb.sheetnames and new_name not in wb.sheetnames:
        new_ws = wb.copy_worksheet(wb["B_temp"])
        new_ws.title = new_name

        # -------- ここがポイント --------
        # base_sheet の直前に挿入する
        try:
            base_pos = wb.sheetnames.index(base_sheet)
        except ValueError:
            base_pos = len(wb.worksheets) - 1      # 念のため: 見つからなければ末尾

        # openpyxl 公式 API（3.0 以降）: move_sheet でも OK
        # wb.move_sheet(new_ws, offset=base_pos - wb.sheetnames.index(new_name))

        # 内部リストを直接操作
        wb._sheets.remove(new_ws)
        wb._sheets.insert(base_pos, new_ws)
        # ---------------------------------

        wb.save(file_path)

    wb.close()


# ---------- Sheet1 を除去するヘルパ ----------
def remove_sheet1(wb):
    """
    ワークブックに 'Sheet1' があれば削除する。
    Excel新規作成時のデフォルトシート対策。
    """
    if "Sheet1" in wb.sheetnames:
        wb.remove(wb["Sheet1"])



def get_pointer(conn: sqlite3.Connection, name: str):
    """
    指定nameのpointer（個人ファイルの書き込み位置）を取得。
    なければNoneを返す。
    """
    row = conn.execute(
        "SELECT file, sheet, next_row FROM personal_pointer WHERE name = ?", (name,)
    ).fetchone()
    return row  # None or (file, sheet, next_row)


def set_pointer(conn: sqlite3.Connection, name: str, file: str, sheet: str, next_row: int):
    """
    pointer情報をINSERTまたはUPDATEする。
    """
    conn.execute('''
        INSERT INTO personal_pointer (name, file, sheet, next_row)
        VALUES (?, ?, ?, ?)
        ON CONFLICT(name)
        DO UPDATE SET file=excluded.file, sheet=excluded.sheet, next_row=excluded.next_row
    ''', (name, file, sheet, next_row))


def increment_sheet_name(base: str, idx: int) -> str:
    """
    base='宮本武蔵', idx=2 の場合 '宮本武蔵(2)' を返す。
    """
    return f"{base}({idx})"


def wareki_year(year: int) -> int:
    """
    西暦年を令和年に変換（2019年=令和1年）。
    """
    return year - 2018

def add_footer(file_path: str, base_sheet: str):
    """
    “○日裏”シリーズの最後尾シートにFooterを貼り付ける。
    ・行37が空ならその行に貼り付け
    ・埋まっていれば新しい裏シートを作成し2行目に貼り付け
    行高は33ptに設定
    """
    wb = openpyxl.load_workbook(file_path)
    tpl_footer = wb["Footer"]

    # 末尾シート名を特定
    base, *tail = base_sheet.split("(")         # '15日裏'
    idx = 1
    while f"{base}({idx})" in wb.sheetnames:
        idx += 1
    last_name = base if idx == 1 else f"{base}({idx-1})"
    ws = wb[last_name]

    def paste(ws_target, dest_row):
        for c in range(1, 3):                   # A,B 列
            src = tpl_footer.cell(1, c)
            dst = ws_target.cell(dest_row, c)
            dst.value         = src.value
            dst.font          = copy(src.font)
            dst.border        = copy(src.border)
            dst.fill          = copy(src.fill)
            dst.number_format = src.number_format
            dst.alignment     = copy(src.alignment)
            dst.protection    = copy(src.protection)
        ws_target.row_dimensions[dest_row].height = 33

    # 行37が空ならそこへ貼り付け
    if not ws.cell(37, 1).value and not ws.cell(37, 2).value:
        paste(ws, 37)
    else:
        # 新しい裏シートを作成
        new_name = f"{base}({idx})"
        if "B_temp" in wb.sheetnames:
            ws_new = wb.copy_worksheet(wb["B_temp"])
            ws_new.title = new_name
        else:
            ws_new = wb.create_sheet(new_name)
        paste(ws_new, 2)        # 2 行目に貼り付け

    wb.save(file_path)
    wb.close()

PREF_FILE = Path().resolve() / "prefs.json"

def load_prefs():
    """
    担当者名や前回日付などの設定をprefs.jsonから読み込む。
    読み込み失敗時は空のデフォルト値を返す。
    """
    if PREF_FILE.exists():
        try:
            with open(PREF_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except json.JSONDecodeError:
            pass
    return {"author_day": "", "author_night": "", "last_date": ""}

def save_prefs(author_day, author_night, last_date=""):
    """
    担当者名や前回日付などの設定をprefs.jsonに保存する。
    """
    with open(PREF_FILE, "w", encoding="utf-8") as f:
        json.dump(
            {
                "author_day":  author_day,
                "author_night": author_night,
                "last_date":   last_date,
            },
            f,
            ensure_ascii=False,
            indent=2,
        )


# ------------------------------------------------------------------
# Part B : transfer_to_personal_files
# ------------------------------------------------------------------

ROW_LIMIT = 31                     # 1 シート 31 行
PERSONAL_TEMPLATE_SHEET = "personal"
PF_2F   = "2階個人ファイル.xlsx"
PF_3F   = "3階個人ファイル.xlsx"
PF_RET  = "退所者個人ファイル.xlsx"

WEEKDAY_STR = "月火水木金土日"


def ensure_personal_file(base_dir: Path, file_name: str, template_src: Path) -> Path:
    """
    個人ファイル（2階/3階/退職者）がなければテンプレートから複製して作成。
    """
    dest = base_dir / file_name
    if not dest.exists():
        shutil.copy(template_src, dest)
    return dest


def ensure_personal_sheet(wb, base_name: str, wareki: int):
    """
    指定名のシートがなければテンプレートから複製して作成し、ヘッダを書き込む。
    戻り値: (sheet_object, 次に書く行番号)
    """
    sheet = wb[base_name] if base_name in wb.sheetnames else None
    new_created = False

    if sheet is None:
        # テンプレ personal を複製 / fallback create
        if PERSONAL_TEMPLATE_SHEET in wb.sheetnames:
            sheet = wb.copy_worksheet(wb[PERSONAL_TEMPLATE_SHEET])
            sheet.title = base_name
        else:
            sheet = wb.create_sheet(base_name)
        sheet["A2"] = f"令和{wareki}年"
        sheet["C2"] = f"　入所者氏名　{base_name}"
        new_created = True
        remove_sheet1(wb)

    # ------- 「次に書く行」を決定 -------
    if new_created:
        next_row = 4                      # 新規なら 4 行目から
    else:
        # 4 行目以降を上から見て、完全に空白の最初の行を探す
        max_row = sheet.max_row
        next_row = None
        for r in range(4, max_row + 1):
            if all(sheet.cell(r, c).value in (None, "") for c in range(1, 5)):
                next_row = r
                break
        if next_row is None:              # どこにも空きが無ければ末尾の次行
            next_row = max_row + 1

    return sheet, next_row



def transfer_to_personal_files(entries: list, date: dt.datetime,
                               db_path: str, base_dir: Path, template_src: Path):
    """
    日誌エントリ（entries）を各入所者の個人ファイル（Excel）に転記する。
    必要に応じて新規シート作成や年切り替え、行数超過時の分割も自動で行う。
    entries: [{name, content, room, shift, author}]
    date: 転記日付
    db_path: ポインタ管理用DBパス
    base_dir: 個人ファイル保存先ディレクトリ
    template_src: テンプレートExcelファイルパス
    """
    init_personal_tables(db_path)  # ポインタテーブルがなければ作成

    yyyy = date.year
    wareki = wareki_year(date.year)

    # 個人ファイルのワークブックキャッシュ
    cache = {}

    conn = sqlite3.connect(db_path)

    for ent in entries:
        name   = ent["name"]
        room   = ent["room"]
        author = ent["author"]
        shift  = ent["shift"]
        md_str = f"{date.month}/{date.day}"
        wday   = WEEKDAY_STR[date.weekday()]
        content = ent["content"]

        # --- どの個人ファイルに書くか判定 ---
        if room == "退所":
            pf_name = PF_RET
        elif room.startswith("2"):
            pf_name = PF_2F
        elif room.startswith("3"):
            pf_name = PF_3F
        else:
            pf_name = PF_RET  # 不明は退職者へ

        # --- 個人ファイル（Excel）を用意 ---
        if pf_name not in cache:
            cache[pf_name] = openpyxl.load_workbook(
                ensure_personal_file(base_dir, pf_name, template_src)
            )
        wb = cache[pf_name]

        # --- どのシート・行に書くかポインタ取得 ---
        ptr = get_pointer(conn, name)
        if ptr and ptr[0] == pf_name and ptr[1] in wb.sheetnames:
            sheet = wb[ptr[1]]
            next_row = ptr[2]
        else:
            sheet, next_row = ensure_personal_sheet(wb, name, wareki)

        # --- シートの行数上限を超える場合は新シート作成 ---
        if next_row > (ROW_LIMIT + 3):
            idx = 2
            while increment_sheet_name(name, idx) in wb.sheetnames:
                idx += 1
            new_title = increment_sheet_name(name, idx)

            new_ws = copy_left_of(wb, sheet, PERSONAL_TEMPLATE_SHEET, new_title)
            new_ws["A2"] = f"令和{wareki}年"
            new_ws["C2"] = f"　入所者氏名　{name}"

            sheet   = new_ws
            next_row = 4

        # --- 年度が変わった場合は区切りを挿入 ---
        current_wareki = re.search(r"令和(\d+)年", str(sheet["B1"].value))
        current_wareki = int(current_wareki.group(1)) if current_wareki else wareki
        if current_wareki != wareki:
            sheet.cell(row=next_row, column=4, value=f"ここから令和{wareki}年")
            next_row += 1
            sheet["A2"] = f"令和{wareki}年"

        # --- 本文を複数行に分割して書き込む ---
        lines = [ln for ln in content.split("\n") if ln]  # 空行は捨てる
        rows_needed = len(lines)

        # --- 残り行が足りない場合は新シート作成 ---
        if next_row + rows_needed - 1 > (ROW_LIMIT + 3):
            idx = 2
            while increment_sheet_name(name, idx) in wb.sheetnames:
                idx += 1
            new_title = increment_sheet_name(name, idx)

            new_ws = copy_left_of(wb, sheet, PERSONAL_TEMPLATE_SHEET, new_title)
            new_ws["A2"] = f"令和{wareki}年"
            new_ws["C2"] = f"　入所者氏名　{name}"

            sheet   = new_ws
            next_row = 4

        # --- 行ごとに日付・曜日・本文・記録者を書き込む ---
        for i, line in enumerate(lines):
            if i == 0:
                # 最初の行だけ日付と曜日を書く
                sheet.cell(next_row, 1, md_str)
                sheet.cell(next_row, 2, wday)
            # 本文（列3）を書き込む
            sheet.cell(next_row, 3, line)
            if i == len(lines) - 1:
                # 最終行だけ記録者を書く
                sheet.cell(next_row, 4, author)
            next_row += 1

        # --- ポインタ情報をDBに保存 ---
        set_pointer(conn, name, pf_name, sheet.title, next_row)

    conn.commit()
    conn.close()

    # --- すべての個人ファイルを保存 ---
    for path_name, wb in cache.items():
        wb.save(base_dir / path_name)

        # --- ROW_LIMIT 超過チェック ---
        if next_row > (ROW_LIMIT + 3):
            idx = 2
            while increment_sheet_name(name, idx) in wb.sheetnames:
                idx += 1
            new_title = increment_sheet_name(name, idx)

            new_ws = copy_left_of(wb, sheet, PERSONAL_TEMPLATE_SHEET, new_title)
            new_ws["A2"] = f"令和{wareki}年"
            new_ws["C2"] = f"　入所者氏名　{name}"

            sheet   = new_ws
            next_row = 4

        # 年が変わった場合（シートの B1 年と比較）
        current_wareki = re.search(r"令和(\d+)年", str(sheet["B1"].value))
        current_wareki = int(current_wareki.group(1)) if current_wareki else wareki
        if current_wareki != wareki:
            sheet.cell(row=next_row, column=4, value=f"ここから令和{wareki}年")
            next_row += 1
            sheet["A2"] = f"令和{wareki}年"

        #-----この記事を書き込む処理（複数行対応）-----
        # content を改行で分割 → ['１行目', '２行目', ...]
        lines = [ln for ln in content.split("\n") if ln]  # 空行は捨てる
        rows_needed = len(lines)

        # --- 残り行が足りないとき ---
        if next_row + rows_needed - 1 > (ROW_LIMIT + 3):
            idx = 2
            while increment_sheet_name(name, idx) in wb.sheetnames:
                idx += 1
            new_title = increment_sheet_name(name, idx)

            new_ws = copy_left_of(wb, sheet, PERSONAL_TEMPLATE_SHEET, new_title)
            new_ws["A2"] = f"令和{wareki}年"
            new_ws["C2"] = f"　入所者氏名　{name}"

            sheet   = new_ws
            next_row = 4

        #------行リストを順番に書き込む------
        for i, line in enumerate(lines):
            if i == 0:
                # 最初の行だけ日付と曜日を書く
                sheet.cell(next_row, 1, md_str)
                sheet.cell(next_row, 2, wday)
            # 本文（列3）を書き込む
            sheet.cell(next_row, 3, line)
            if i == len(lines) - 1:
                # 最終行だけ記録者を書く
                sheet.cell(next_row, 4, author)
            next_row += 1


        # pointer 更新
        set_pointer(conn, name, pf_name, sheet.title, next_row)

    conn.commit()
    conn.close()

    # ブック保存
    for path_name, wb in cache.items():
        wb.save(base_dir / path_name)


def normalize_text(text: str) -> str:
    """
    全角→半角変換や空白除去など、テキストを正規化して返す。
    """
    if not text:
        return ""
    text = text.translate(str.maketrans({
        '　': ' ',
        '０': '0', '１': '1', '２': '2', '３': '3', '４': '4',
        '５': '5', '６': '6', '７': '7', '８': '8', '９': '9'
    }))
    return text.strip()


def create_database_if_not_exists(db_path: str):
    """
    residents/diary_entriesテーブルがなければ作成する。
    DB初期化用。
    """
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS residents (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            room TEXT NOT NULL,
            birthday TEXT,
            gender TEXT
        )
    ''')
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS diary_entries (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            resident_name TEXT NOT NULL,
            date TEXT NOT NULL,
            shift TEXT CHECK(shift IN ('日勤', '夜勤')),
            content TEXT,
            author TEXT
        )
    ''')
    cursor.execute('''
        CREATE UNIQUE INDEX IF NOT EXISTS uniq_entry
        ON diary_entries(resident_name, date, shift, content)
    ''')   
    conn.commit()
    conn.close()

# -----------------------------------------------------------------------------
#  ユーティリティ
# -----------------------------------------------------------------------------

def normalize_text(val: str) -> str:
    """None → 空文字 / 前後空白除去。セル値を安定化。"""
    return val.strip() if val else ""


def iter_rows(sheet, start: int = 2) -> Iterator[tuple[int, str, str]]:
    """指定行から A 列・B 列の文字列を順に返す。"""
    row = start
    while row <= sheet.max_row:
        name = sheet.cell(row, 1).value
        content = sheet.cell(row, 2).value
        yield row, normalize_text(str(name)) if name else "", normalize_text(str(content)) if content else ""
        row += 1


def apply_night_header(sheet, row: int, template_sheet) -> None:
    """`template_sheet` 先頭 2 行を `row` 位置へコピー。"""
    for offset in range(HEADER_ROWS):
        for col in range(1, 3):  # A 列・B 列だけコピー（必要なら拡張）
            src_cell = template_sheet.cell(offset + 1, col)
            dst_cell = sheet.cell(row + offset, col)
            dst_cell.value = src_cell.value
            if src_cell.has_style:
                dst_cell._style = src_cell._style


def setup_page_breaks(sheet, rows_per_page: int = ARTICLE_ROWS_PER_PAGE) -> None:
    """長大シートを印刷する際の手動改ページを自動挿入する。"""
    from openpyxl.worksheet.pagebreak import Break

    # 既存ブレークをリセット
    sheet.page_breaks = []

    total_rows = sheet.max_row
    idx = rows_per_page + 1  # 改ページ位置 = 次ページ先頭行
    while idx < total_rows:
        sheet.page_breaks.append(Break(id=idx))
        idx += rows_per_page

# -----------------------------------------------------------------------------
#  メイン関数
# -----------------------------------------------------------------------------


    
def create_input_sheet(template_path: str, target_path: str, date: dt):
    if not Path(target_path).exists():
        shutil.copy(template_path, target_path)

    wb = openpyxl.load_workbook(target_path)
    remove_sheet1(wb)
    weekday = "月火水木金土日"[date.weekday()]
    wareki = wareki_year(date.year) 
    sheet表 = f"{date.day}日表"
    sheet裏 = f"{date.day}日裏"

    # ---- 表シート ----
    if "F_temp" in wb.sheetnames and sheet表 not in wb.sheetnames:
        new_ws = wb.copy_worksheet(wb["F_temp"])
        new_ws.title = sheet表

        # ★ ここで先頭へ移動 ★
        wb._sheets.remove(new_ws)     # 末尾から一旦削除
        wb._sheets.insert(0, new_ws)  # インデックス 0 (左端) へ

        new_ws["A2"] = f"令和{wareki_year(date.year)}年"
        new_ws["A3"] = f"{date.month}月{date.day}日（{weekday}) 天気"

    # ---- 裏シート ----
    if "B_temp" in wb.sheetnames and sheet裏 not in wb.sheetnames:
        new_ws = wb.copy_worksheet(wb["B_temp"])
        new_ws.title = sheet裏
        wb._sheets.remove(new_ws)
        wb._sheets.insert(0, new_ws)   # 表と同じく左端へ

    wb.save(target_path)
    os.startfile(target_path)


def add_ura_sheet(template_path: str, target_path: str, date: dt):
    sheet_base = f"{date.day}日裏"
    wb = openpyxl.load_workbook(target_path)
    remove_sheet1(wb)

    if sheet_base not in wb.sheetnames:
        messagebox.showerror("エラー", f"{sheet_base} が存在しません。先に日誌を作成してください。")
        return

    # 新しい裏番号を決定（例: 〇日裏(2), (3), ...）
    i = 2
    while True:
        new_sheet_name = f"{sheet_base}({i})"
        if new_sheet_name not in wb.sheetnames:
            break
        i += 1

    if "B_temp" not in wb.sheetnames:
        messagebox.showerror("エラー", "テンプレート B_temp が見つかりません。")
        return

    new_sheet = wb.copy_worksheet(wb["B_temp"])
    wb._sheets.remove(new_sheet)  # 末尾に追加されるので先に削除
    base_index = wb.sheetnames.index(sheet_base)
    wb._sheets.insert(base_index, new_sheet)  # 指定位置に挿入
    new_sheet.title = new_sheet_name

    wb.save(target_path)
    os.startfile(target_path)

ROOM_SEQ = [str(i) for i in range(201, 226)] + [str(i) for i in range(301, 326)]

def update_resident(name, room, birthday, gender, db_path, excel_path):
    # ---------- DB ----------
    conn = sqlite3.connect(db_path)
    cur  = conn.cursor()

    cur.execute("SELECT 1 FROM residents WHERE name = ?", (name,))
    if cur.fetchone():
        cur.execute("""UPDATE residents
                       SET room=?, birthday=?, gender=?
                       WHERE name=?""",
                    (room, birthday, gender, name))
    else:
        cur.execute("""SELECT name FROM residents
                       WHERE room=? AND room NOT IN ('退所','保留')""", (room,))
        dup = cur.fetchone()
        if dup:
            cur.execute("UPDATE residents SET room='保留' WHERE name=?",
                        (dup[0],))
            messagebox.showinfo("居室重複",
                                f"{dup[0]} さんの居室番号を保留としています")
        cur.execute("""INSERT INTO residents
                       (name, room, birthday, gender)
                       VALUES (?,?,?,?)""",
                    (name, room, birthday, gender))

    conn.commit()

    # ---------- データ取得 ----------
    cur.execute("""SELECT name, room, birthday, gender
                   FROM residents WHERE room!='退所'""")
    rows = cur.fetchall()
    conn.close()

    by_room = {r[1]: r for r in rows if r[1] in ROOM_SEQ}
    on_hold = sorted((r for r in rows if r[1] == '保留'),
                     key=lambda x: x[0])         # 名前順

    # ---------- Excel ----------
    wb = openpyxl.load_workbook(excel_path)
    if "入所者名簿" not in wb.sheetnames:
        wb.create_sheet("入所者名簿")
    ws = wb["入所者名簿"]

    ws.delete_rows(1, ws.max_row)                # 全削除

    # A1 に更新日（JST）
    ws['A1'] = dt.datetime.now(
                  dt.timezone(dt.timedelta(hours=9))
               ).strftime('%Y年%m月%d日　更新')


    # 固定順の 50 行（201-225, 301-325）
    for rm in ROOM_SEQ:
        if rm in by_room:
            name, _, birth, sex = by_room[rm]
            # append は現在の最終行+1 から入るので A1 の直後=行2 から並ぶ
            if rm.isdigit():
                ws.append([name, int(rm), birth, sex])
            else:                                # 万一文字付き番号
                ws.append([name, rm, birth, sex])
                ws.cell(row=ws.max_row, column=2).number_format = "@"
        else:
            ws.append(["", "", "", ""])          # 空室

    # '保留' を最後に
    for name, rm, birth, sex in on_hold:
        ws.append([name, rm, birth, sex])
        ws.cell(row=ws.max_row, column=2).number_format = "@"

    wb.save(excel_path)

def manage_residents_ui(db_path, excel_path):
    win = tk.Toplevel()
    win.title("入所者名簿管理")

    tk.Label(win, text="氏名").grid(row=0, column=0)
    name_entry = tk.Entry(win)
    name_entry.grid(row=0, column=1)

    tk.Label(win, text="居室番号").grid(row=1, column=0)
    room_entry = tk.Entry(win)
    room_entry.grid(row=1, column=1)

    tk.Label(win, text="生年月日").grid(row=2, column=0)
    birth_y = tk.Entry(win, width=6)
    birth_y.grid(row=2, column=1, sticky="w")
    tk.Label(win, text="月").grid(row=2, column=1, padx=(40, 0))

    birth_m = tk.Entry(win, width=4)
    birth_m.grid(row=2, column=1, padx=(80, 0), sticky="w")
    tk.Label(win, text="日").grid(row=2, column=1, padx=(110, 0))

    birth_d = tk.Entry(win, width=4)
    birth_d.grid(row=2, column=1, padx=(140, 0), sticky="w")
    tk.Label(win, text="日").grid(row=2, column=1, padx=(170, 0))

    gender_var = tk.StringVar(value="男性")
    tk.Radiobutton(win, text="男性", variable=gender_var, value="男性").grid(row=3, column=0)
    tk.Radiobutton(win, text="女性", variable=gender_var, value="女性").grid(row=3, column=1)

    def register():
        name = name_entry.get().strip()
        room = room_entry.get().strip()
        y, m, d = birth_y.get(), birth_m.get(), birth_d.get()
        try:
            birthday = f"{int(y):04d}-{int(m):02d}-{int(d):02d}"
        except:
            messagebox.showerror("エラー", "生年月日が正しくありません")
            return
        gender = gender_var.get()
        if not name or not room:
            messagebox.showerror("エラー", "氏名と居室番号を入力してください")
            return
        update_resident(name, room, birthday, gender, db_path, excel_path)
        messagebox.showinfo("完了", "登録が完了しました。")

    tk.Button(win, text="新規登録", command=register).grid(row=4, column=0, columnspan=2, pady=10)

    # ---------------------------------------------------------------------------
    #  個人ファイル転記メインフロー (GUI から呼ばれる想定)
    # ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------
#  個人ファイル転記メインフロー (GUI から呼ばれる想定)
# ---------------------------------------------------------------------------

def personal_transfer(
    date: dt.date,
    author_day: str,
    author_night: str,
    base_dir: Path,
    template_xlsx: Path,
):
    """GUI 側で呼び出す転記エントリポイント。"""

    if not author_day or not author_night:
        messagebox.showerror("エラー", "日勤と夜勤の担当者名を入力してください。")
        return

    yyyy, mm = date.year, date.month
    target_file = base_dir / f"{yyyy}_{mm:02d}_処遇日誌.xlsx"
    db_path     = base_dir / f"diary_{yyyy}.db"

    # shutil.copyfile の方がメモリ効率◎
    if not target_file.exists():
        target_file.write_bytes(template_xlsx.read_bytes())

    wb = openpyxl.load_workbook(target_file)
    sheet_name = f"{date.day}日裏"
    if sheet_name not in wb.sheetnames:
        messagebox.showerror("エラー", f"シート {sheet_name} が見つかりません")
        wb.close()
        return

    sheet = wb[sheet_name]
    night_tpl = wb["Header_Night"] if "Header_Night" in wb.sheetnames else None

    entries = extract_entries(sheet)
    if not entries:
        messagebox.showinfo("確認", "転記対象の記事がありません。")
        wb.close()
        return

    entries = add_authors(entries, author_day=author_day, author_night=author_night)

    save_entries_to_db(entries, db_path, date=date)  # DB スキーマ存在確認必須

    update_diary_sheet(sheet, template_sheet=night_tpl)
    wb.save(target_file)
    wb.close()

    # --- 個人ファイル転記 ---
    if transfer_to_personal_files:
        transfer_to_personal_files(entries, date, db_path, base_dir, template_xlsx)
    else:
        messagebox.showwarning("警告", "transfer_to_personal_files が見つかりません")

    # --- 夜勤フッター ---
    if any(e["shift"] == "夜勤" for e in entries):
        if add_footer:
            add_footer(str(target_file), sheet_name)
        else:
            messagebox.showwarning("警告", "add_footer が見つかりません")

    messagebox.showinfo("完了", "個人ファイルへの転記と DB 登録が完了しました。")



def main_ui():
    """
    メインのTkinter GUI画面を表示し、各種操作（作成・転記・名簿管理など）を行う。
    """
    root = tk.Tk()
    root.title("処遇日誌アプリ")
    root.geometry("300x500")


    prefs = load_prefs()                 # ← ここで読込
    today  = dt.datetime.now()

    # -------- 年 --------
    tk.Label(root, text="西暦:", font=("Arial", 14)).grid(row=0, column=0, padx=10, pady=10)
    year_entry = tk.Entry(root, font=("Arial", 14), width=8)
    year_entry.grid(row=0, column=1)

    # -------- 月 --------
    tk.Label(root, text="月:", font=("Arial", 14)).grid(row=1, column=0, padx=10, pady=10)
    month_entry = tk.Entry(root, font=("Arial", 14), width=8)
    month_entry.grid(row=1, column=1)

    # -------- 日 --------
    tk.Label(root, text="日:", font=("Arial", 14)).grid(row=2, column=0, padx=10, pady=10)
    day_entry = tk.Entry(root, font=("Arial", 14), width=8)
    day_entry.grid(row=2, column=1)

    # ==== 既定値を入れる ====
    if prefs.get("last_date", "") and prefs["last_date"] != today.strftime("%Y-%m-%d"):
        # 'YYYY-MM-DD' → 年月日に分解
        y, m, d = map(int, prefs["last_date"].split("-"))
    else:
        y, m, d = today.year, today.month, today.day

    year_entry.insert(0, y)
    month_entry.insert(0, m)
    day_entry.insert(0, d)

    # システム日付と異なっていればソフトに警告
    if prefs.get("last_date", "") and prefs.get("last_date", "") != today.strftime("%Y-%m-%d"):
        messagebox.showinfo(
            "確認",
            f"前回作業日は {prefs['last_date']} です。\n"
            f"PCの日付は {today.strftime('%Y-%m-%d')} になっています。",
        )
    def get_date():
        try:
            return dt.datetime(int(year_entry.get()), int(month_entry.get()), int(day_entry.get()))
        except ValueError:
            messagebox.showerror("エラー", "正しい日付を入力してください。")
            return None

    def make_input_sheet():
        date = get_date()
        if not date:
            return
        yyyy, mm = date.year, date.month
        target_file = Path().resolve() / f"{yyyy}_{mm:02d}_処遇日誌.xlsx"
        template = Path().resolve() / "Tre_diary_temp.xlsx"
        if not target_file.exists():
            shutil.copy(template, target_file)
        create_input_sheet(str(template), str(target_file), date)


    def add_extra_ura():
        date = get_date()
        if not date:
            return
        yyyy = int(year_entry.get())
        mm = int(month_entry.get())
        target_file = Path().resolve() / f"{yyyy}_{mm:02d}_処遇日誌.xlsx"
        template = Path().resolve() / "Tre_diary_temp.xlsx"
        if not target_file.exists():
            shutil.copy(template, target_file)
        add_ura_sheet(str(template), str(target_file), date)



    def open_resident_manager():
        base = Path().resolve()
        db_file = base / "diary_2025.db"
        excel_file = base / "入所者名簿.xlsx"
        create_database_if_not_exists(str(db_file))
        if not excel_file.exists():
            messagebox.showerror("エラー", "入所者名簿ファイルが見つかりません。")
            return
        manage_residents_ui(str(db_file), str(excel_file))

    
    def save_authors():
        save_prefs(
            author_day_var.get().strip(),
            author_night_var.get().strip(),
            # last_date は「現在 UI に表示されている日付」
            f"{year_entry.get()}-{int(month_entry.get()):02d}-{int(day_entry.get()):02d}",
        )
        messagebox.showinfo("保存", "担当者名と日付を保存しました。")


    # 日勤担当者欄
    tk.Label(root, text="日勤担当者:", font=("Arial", 12)).grid(row=3, column=0, sticky="e")
    author_day_var = tk.StringVar(value=prefs["author_day"])   # デフォルトセット
    tk.Entry(root, textvariable=author_day_var, width=12).grid(row=3, column=1, sticky="w")

    # 夜勤担当者欄
    tk.Label(root, text="夜勤担当者:", font=("Arial", 12)).grid(row=4, column=0, sticky="e")
    author_night_var = tk.StringVar(value=prefs["author_night"])
    tk.Entry(root, textvariable=author_night_var, width=12).grid(row=4, column=1, sticky="w")
    
    tk.Button(root, text="担当者を保存", command=save_authors)\
        .grid(row=5, column=0, columnspan=2, pady=5)

    tk.Button(root, text="日誌作成", font=("Arial", 14), command=make_input_sheet).grid(row=6, column=0, columnspan=2, pady=10)
    tk.Button(root, text="日誌裏追加", font=("Arial", 14), command=add_extra_ura).grid(row=7, column=0, columnspan=2, pady=10)
    tk.Button(root, text="個人ファイルに転記",
          font=("Arial", 14),
          command=personal_transfer).grid(row=8, column=0, columnspan=2, pady=10)

    tk.Button(root, text="入所者名簿管理", font=("Arial", 14), command=open_resident_manager).grid(row=9, column=0, columnspan=2, pady=10)

    root.mainloop()


if __name__ == "__main__":
    # 今年の DB パスを決めてテーブルを保証
    db_file = Path().resolve() / f"diary_{dt.datetime.now().year}.db"
    init_personal_tables(str(db_file))

    # メイン画面
    main_ui()

